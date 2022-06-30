from bs4 import BeautifulSoup
import requests
import openpyxl
excel=openpyxl.Workbook()
#print(excel.sheetnames)
sheet=excel.active # Active sheet where we load our data
sheet.title="Top Rated Movies"
#print(excel.sheetnames)
sheet.append(["Movie_Rank","Movie_Name","Year_of_Release","IMDB_Rating"])

try:
    source=requests.get("https://www.imdb.com/chart/top/") # To access site
    source.raise_for_status() # Throw an error if above link is niot exits

    soup=BeautifulSoup(source.text,'html.parser') # It returns the html code

    movies=soup.find('tbody',class_="lister-list").find_all('tr')

    for movie in movies:
        name=movie.find('td',class_="titleColumn").a.text

        rank=movie.find('td',class_="titleColumn").get_text(strip=True).split(".")[0]

        year=movie.find('td',class_="titleColumn").span.text.strip("()")

        rating=movie.find('td',class_="ratingColumn imdbRating").strong.text
        #print(rank,name,year,rating)
        sheet.append([rank,name,year,rating])
        
except Exception as e:
    print(e)
excel.save('D:\Py_Practice\IMDB_Movie_Rating.xlsx')